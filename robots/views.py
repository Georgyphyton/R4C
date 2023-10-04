from django.views import View
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from datetime import date, timedelta
from openpyxl import Workbook
from .models import Robot
from .forms import ValidationForm
import json


def index(request):
    return render(
        request,
        'robots/index.html',
    )


@method_decorator(csrf_exempt, name='dispatch')
class PostRobotView(View):
    def post(self, request):
        post_body = json.loads(request.body)
        form = ValidationForm(post_body)
        if not form.is_valid():
            data = {
                'message': 'Incorrect data entry format!',
            }
            return JsonResponse(data, status=400)
        robot_version = post_body.get('version')
        robot_model = post_body.get('model')
        data = post_body.get('created')
        robot = {
            'model': robot_model,
            'version': robot_version,
            'serial': f'{robot_model}-{robot_version}',
            'created': data,
        }
        robot_obj = Robot.objects.create(**robot)
        data = {
            'message': f'New robot has been created with id {robot_obj.id}'
        }
        return JsonResponse(data, status=201)
    
    def export_to_excel(request):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="robots.xlsx"'
        wb = Workbook()
        headers = ["Модель", "Версия", "Количество за неделю"]
        today = date.today()
        seven_day_before = today - timedelta(days=7)
        robots = Robot.objects.filter(created__gte=seven_day_before).values('model', 'version').annotate(created_count=Count('id')).order_by('model', 'version')
        for robot in robots:
            if robot['model'] not in wb.sheetnames:
                wb.create_sheet(title=robot['model'])
                wb[robot['model']].append(headers)
            wb[robot['model']].append([robot['model'], robot['version'], robot['created_count']])
        if len(wb.sheetnames) > 1:
            wb.remove(wb['Sheet'])
        wb.save(response)
        return response
